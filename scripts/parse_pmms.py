"""Extract 15-yr and 30-yr FRM weekly rates from Freddie Mac PMMS,
compute monthly averages for June 2024 - May 2026, emit two JSON files."""
import openpyxl
from collections import defaultdict
from datetime import datetime
import json
import sys

WB_PATH = r"C:\Users\akanc\Documents\freddie_pmms.xlsx"
OUT_DIR = r"c:\GitHub\Mortgage_Loan_Dashboard\src\data"

# PMMS "Full History" layout (header rows 5-7 form a stacked label):
# Row 5: B=U.S., C=30 yr, D=U.S., E=15 yr
# Row 6: B="30 yr", C="fees &", D="15 yr", E="fees &"
# Row 7: B=FRM, C=points, D=FRM, E=points
# Data starts row 8. Column A=Week date, B=30-yr FRM, D=15-yr FRM.
COL_DATE = 1
COL_30YR = 2
COL_15YR = 4

wb = openpyxl.load_workbook(WB_PATH, data_only=True)
ws = wb["Full History"]

# --- Header verification: row 6 holds the term labels ("30 yr" / "15 yr") ---
header_row6 = [ws.cell(row=6, column=c).value for c in range(1, 6)]
print(f"Row 6 (term labels): {header_row6}")
h30 = str(header_row6[COL_30YR - 1] or "")
h15 = str(header_row6[COL_15YR - 1] or "")
if "30" not in h30:
    sys.exit(f"ERROR: expected '30' in row-6 column B, got: {h30!r}")
if "15" not in h15:
    sys.exit(f"ERROR: expected '15' in row-6 column D, got: {h15!r}")
print(f"Headers OK: col B = {h30!r}, col D = {h15!r}\n")

monthly_15 = defaultdict(list)
monthly_30 = defaultdict(list)

for r in range(8, ws.max_row + 1):
    wk = ws.cell(row=r, column=COL_DATE).value
    if not isinstance(wk, datetime):
        continue
    # Window: June 2024 through May 2026
    in_window = (wk.year == 2024 and wk.month >= 6) or wk.year == 2025 or (wk.year == 2026 and wk.month <= 5)
    if not in_window:
        continue
    key = (wk.year, wk.month)
    rate_15 = ws.cell(row=r, column=COL_15YR).value
    rate_30 = ws.cell(row=r, column=COL_30YR).value
    if isinstance(rate_15, (int, float)):
        monthly_15[key].append((wk.date().isoformat(), float(rate_15)))
    if isinstance(rate_30, (int, float)):
        monthly_30[key].append((wk.date().isoformat(), float(rate_30)))


def summarize(monthly, label):
    print(f"=== {label} ===")
    print(f"{'Month':<10} {'Avg %':<8} {'N weeks':<8}")
    print("-" * 30)
    results = []
    for ym in sorted(monthly.keys()):
        weeks = monthly[ym]
        avg = sum(r for _, r in weeks) / len(weeks)
        m_label = f"{ym[0]}-{ym[1]:02d}"
        print(f"{m_label:<10} {avg:<8.3f} {len(weeks):<8}")
        results.append({"month": m_label, "rate": round(avg, 3), "n_weeks": len(weeks)})
    return results


results_15 = summarize(monthly_15, "15-yr FRM monthly")
print()
results_30 = summarize(monthly_30, "30-yr FRM monthly")

with open(f"{OUT_DIR}\\pmms_15yr_monthly.json", "w") as f:
    json.dump(results_15, f, indent=2)
with open(f"{OUT_DIR}\\pmms_30yr_monthly.json", "w") as f:
    json.dump(results_30, f, indent=2)
print(f"\nSaved {len(results_15)} months -> pmms_15yr_monthly.json")
print(f"Saved {len(results_30)} months -> pmms_30yr_monthly.json")
